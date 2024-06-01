import openpyxl
import random
from openpyxl.utils import get_column_letter

#menghasilkan nama random
def generate_random_name():
    first_names = ["Adi", "Budi", "Citra", "Dian", "Eka", "Fitri", "Gita", "Hadi", "Indra", "Joko"]
    last_names = ["Wijaya", "Santoso", "Kusuma", "Setiawan", "Pratiwi", "Purnama", "Utami", "Susanto", "Wahyudi", "Widodo"]
    first_name = random.choice(first_names)
    last_name = random.choice(last_names)
    return f"{first_name} {last_name}"

#menghasilkan email random
def generate_random_email(name):
    name = name.lower().replace(" ", "")
    domains = ["gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "mail.com"]
    domain = random.choice(domains)
    return f"{name}@{domain}"

#menghasilkan petugas atau anggota random
def generate_random_status():
    statuses = ["Petugas", "Anggota"]
    return random.choice(statuses)

#menghasilkan password random
def generate_random_password():
    chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890!@#$%^&*"
    password = ""
    for _ in range(8):
        password += random.choice(chars)
    return password

#menghasilkan file excel yang berisi dataset pengguna
def generate_excel_file(filename, num_records):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    #menulis judul atau header
    headers = ["ID", "Nama Lengkap", "Email", "Status", "Password"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header

    #menulis data
    for row in range(2, num_records + 2):
        id_num = row - 1
        nama = generate_random_name()
        email = generate_random_email(nama)
        status = generate_random_status()
        password = generate_random_password()

        worksheet.cell(row=row, column=1, value=id_num)
        worksheet.cell(row=row, column=2, value=nama)
        worksheet.cell(row=row, column=3, value=email)
        worksheet.cell(row=row, column=4, value=status)
        worksheet.cell(row=row, column=5, value=password)

    workbook.save(filename)
    print(f"File {filename} berhasil dibuat.")

#pembuatan file bernama data yang direname menjadi DatasetPengguna
generate_excel_file("data.xlsx", 10)
