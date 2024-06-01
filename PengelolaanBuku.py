import openpyxl
from Backend.Structs.Buku import Buku
from collections import deque

class PengelolaanBuku:      
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet
        self.signature = [] 
        self.books = [] 
        self.read()

    #membaca data buku    
    def read(self):
        sheet = openpyxl.load_workbook(self.file)[self.sheet]
        inc = 0

        for row in sheet.iter_rows():
            inc += 1
            if inc > 5:
                if row[0].value is None:
                    break
                val = [""] * 20
                inc = 0
                for cell in row:
                    if cell.value is None:
                        break
                    val[inc] = str(cell.value)
                    inc += 1
                
                self.books.append(Buku(val[1], val[2], int(val[3]), val[4], val[5], int(val[6]), val[7], int(val[8])))
            else:
                buffer = []
                for cell in row:
                    if cell.value is None:
                        break
                    buffer.append(cell.value)
                self.signature.append(buffer)
       

    #menulis ulang data buku
    def rewrite(self):
        workbook = openpyxl.load_workbook(self.file)
        worksheet = workbook[self.sheet]

        #Delete except the signature
        worksheet.delete_rows(len(self.signature) + 1, worksheet.max_row)

        i = 0
        for book in self.books:
            worksheet.append([i+1] + (book.get_data()))
            i += 1
        

        workbook.save(self.file)
        workbook.close()

    #menambahkan data buku
    def add(self,book):
        self.books.append(book)
        self.rewrite()

    #memperbarui data buku
    def update(self,search_atribute, search_key, update_atribute, update_parameter):
        #search_atribute: pencarian berdasarkan nama dan isbn buku
        #search key: isi variabel dari book.search_atribute
        #update_atribute: atribut yang ingin diupdate pada kolom tersebut
        #update_parameter: parameter yang akan mengubah variabel book.update_atribute

        if(search_atribute != "Judul" and search_atribute != "ISBN"):
            print("Salah atribut!")
            return
    
        for book in self.books:
            if(getattr(book,search_atribute.lower()) == search_key):
                setattr(book,update_atribute,update_parameter)
            
            if(book.jumlah_buku == 0):
                self.books.remove(book)

        self.rewrite()
    
    #menghapus data buku
    def delete(self,search_atribute, search_key):
        if(search_atribute != "Judul" and search_atribute != "ISBN"):
            print("Salah atribut!")
            return

        for book in self.books:
            if(getattr(book,search_atribute.lower()) == search_key):
                self.books.remove(book)
        self.rewrite()
        
    #pencarian buku menggunakan BFS
    def search(self, search_attribute, search_key):
        queue = deque()
        
        #menggunakan binary heap untuk membangun queue awal
        for book in self.books:
            if getattr(book, search_attribute.lower()) == search_key:
                queue.append(book)
        
        #lakukan pencarian dengan queue
        while queue:
            book = queue.popleft()
            
            #mengembalikan nilai jika key nya cocok
            if getattr(book, search_attribute.lower()) == search_key:
                return book
            
        
            if hasattr(book, "books"):
                for nested_book in book.books:
                    if getattr(nested_book, search_attribute.lower()) == search_key:
                        queue.append(nested_book)
        
        return None
